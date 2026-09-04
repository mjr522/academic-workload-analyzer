"""
Department Starter Roster Generator
-----------------------------------
Outputs pre-populated starter roster spreadsheets for Department Heads.
Generates native Excel (.xlsx) workbooks with embedded dropdown validation and
backwards-compatible CSV files.

Pre-fills verified teaching and advising data while providing validated columns for:
- Billet Occupancy: Filled, Vacant
- Billet Type: Military, Civilian, MOA_Courtesy
- Expected Tier: Line_Faculty, Course_Director, Department_Head, Division_Chief, Lab_Director, Research_Exempt
- 4-Bucket FTE Distribution (% Academics, % Admin, % Research, % Lab Ops) with automated sum check
"""

import csv
import os
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class RosterGenerator:
    def __init__(self, metrics_data: Dict[str, Any], output_dir: str = "./rosters"):
        self.metrics_data = metrics_data
        self.output_dir = output_dir

    def generate_all_rosters(self) -> Dict[str, str]:
        """Generates starter roster XLSX and CSV files for each department in the dataset."""
        os.makedirs(self.output_dir, exist_ok=True)
        generated_files = {}

        faculty_list = self.metrics_data.get('faculty_directory', [])
        departments = self.metrics_data.get('departments', [])

        fieldnames = [
            'Department_Code',
            'Faculty_Name',
            'Billet_Occupancy',
            'Billet_Type',
            'Expected_Tier',
            'Assigned_Courses',
            'Weighted_Sections_Taught',
            'Cadet_Contact_Load',
            'Cadet_Advisees_Count',
            'FTE_Academics_Pct',
            'FTE_Admin_Pct',
            'FTE_Research_Pct',
            'FTE_LabOps_Pct',
            'Total_FTE_Pct',
            'Department_Head_Notes'
        ]

        for dept in departments:
            dept_code = dept['dept_code']
            dept_name = dept['dept_name']
            dept_faculty = [f for f in faculty_list if f['primary_dept'] == dept_code]

            # -------------------------------------------------------------
            # 1. Build row data
            # -------------------------------------------------------------
            rows_data = []
            for f in dept_faculty:
                inst_name = f['instructor']
                courses_str = '; '.join(f.get('courses_taught', []))
                secs = f.get('weighted_sections', 0.0)
                cadets = f.get('cadet_load_allocated', 0.0)
                advisees = f.get('advisees_count', 0)

                # Heuristics for initial baseline estimates
                if 'Head' in inst_name or secs <= 1.0:
                    default_tier = 'Department_Head'
                    acad_pct = 0.25
                    admin_pct = 0.70
                    res_pct = 0.05
                    lab_pct = 0.00
                elif secs <= 2.0:
                    default_tier = 'Course_Director'
                    acad_pct = 0.60
                    admin_pct = 0.25
                    res_pct = 0.15
                    lab_pct = 0.00
                else:
                    default_tier = 'Line_Faculty'
                    acad_pct = 0.75
                    admin_pct = 0.10
                    res_pct = 0.15
                    lab_pct = 0.00

                rows_data.append({
                    'Department_Code': dept_code,
                    'Faculty_Name': inst_name,
                    'Billet_Occupancy': 'Filled',
                    'Billet_Type': 'Military',
                    'Expected_Tier': default_tier,
                    'Assigned_Courses': courses_str,
                    'Weighted_Sections_Taught': secs,
                    'Cadet_Contact_Load': cadets,
                    'Cadet_Advisees_Count': advisees,
                    'FTE_Academics_Pct': acad_pct,
                    'FTE_Admin_Pct': admin_pct,
                    'FTE_Research_Pct': res_pct,
                    'FTE_LabOps_Pct': lab_pct,
                    'Total_FTE_Pct': None,
                    'Department_Head_Notes': ''
                })

            # Add sample vacant line row for the DH to calibrate
            rows_data.append({
                'Department_Code': dept_code,
                'Faculty_Name': '[VACANT BILLET EXAMPLE]',
                'Billet_Occupancy': 'Vacant',
                'Billet_Type': 'Civilian',
                'Expected_Tier': 'Line_Faculty',
                'Assigned_Courses': '(Unfilled)',
                'Weighted_Sections_Taught': 0.0,
                'Cadet_Contact_Load': 0.0,
                'Cadet_Advisees_Count': 0,
                'FTE_Academics_Pct': 0.75,
                'FTE_Admin_Pct': 0.10,
                'FTE_Research_Pct': 0.15,
                'FTE_LabOps_Pct': 0.00,
                'Total_FTE_Pct': None,
                'Department_Head_Notes': 'Add vacant or double-billeted lines here'
            })

            # -------------------------------------------------------------
            # 2. Write CSV (Backwards compatibility)
            # -------------------------------------------------------------
            csv_path = os.path.join(self.output_dir, f"starter_roster_{dept_code}.csv")
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                for r in rows_data:
                    row_copy = dict(r)
                    if row_copy['Total_FTE_Pct'] is None:
                        row_copy['Total_FTE_Pct'] = 1.0
                    writer.writerow(row_copy)

            # -------------------------------------------------------------
            # 3. Generate Native Excel (.xlsx) with Dropdown Validation
            # -------------------------------------------------------------
            xlsx_path = os.path.join(self.output_dir, f"starter_roster_{dept_code}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Roster_{dept_code}"
            ws.views.sheetView[0].showGridLines = True

            # Write header row
            ws.append(fieldnames)

            # Styling definitions
            navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            data_font = Font(name="Segoe UI", size=9.5)
            bold_font = Font(name="Segoe UI", size=9.5, bold=True)
            gray_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            for col_idx in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = navy_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.row_dimensions[1].height = 28

            # Write data rows
            max_row = len(rows_data) + 1
            for row_idx, r in enumerate(rows_data, start=2):
                ws.row_dimensions[row_idx].height = 20
                is_even = (row_idx % 2 == 0)

                # Col A: Dept Code
                c_dept = ws.cell(row=row_idx, column=1, value=r['Department_Code'])
                c_dept.alignment = Alignment(horizontal="center")

                # Col B: Faculty Name
                c_name = ws.cell(row=row_idx, column=2, value=r['Faculty_Name'])
                c_name.font = bold_font if not r['Faculty_Name'].startswith('[VACANT') else data_font

                # Col C: Billet_Occupancy
                c_occ = ws.cell(row=row_idx, column=3, value=r['Billet_Occupancy'])
                c_occ.alignment = Alignment(horizontal="center")

                # Col D: Billet_Type
                c_typ = ws.cell(row=row_idx, column=4, value=r['Billet_Type'])
                c_typ.alignment = Alignment(horizontal="center")

                # Col E: Expected_Tier
                c_tier = ws.cell(row=row_idx, column=5, value=r['Expected_Tier'])
                c_tier.alignment = Alignment(horizontal="center")

                # Col F: Assigned_Courses
                ws.cell(row=row_idx, column=6, value=r['Assigned_Courses'])

                # Col G: Weighted_Sections_Taught
                c_sec = ws.cell(row=row_idx, column=7, value=r['Weighted_Sections_Taught'])
                c_sec.number_format = '0.00'
                c_sec.alignment = Alignment(horizontal="right")

                # Col H: Cadet_Contact_Load
                c_cad = ws.cell(row=row_idx, column=8, value=r['Cadet_Contact_Load'])
                c_cad.number_format = '0.0'
                c_cad.alignment = Alignment(horizontal="right")

                # Col I: Cadet_Advisees_Count
                c_adv = ws.cell(row=row_idx, column=9, value=r['Cadet_Advisees_Count'])
                c_adv.number_format = '#,##0'
                c_adv.alignment = Alignment(horizontal="right")

                # Col J: FTE_Academics_Pct
                c_fa = ws.cell(row=row_idx, column=10, value=r['FTE_Academics_Pct'])
                c_fa.number_format = '0%'
                c_fa.alignment = Alignment(horizontal="right")

                # Col K: FTE_Admin_Pct
                c_fad = ws.cell(row=row_idx, column=11, value=r['FTE_Admin_Pct'])
                c_fad.number_format = '0%'
                c_fad.alignment = Alignment(horizontal="right")

                # Col L: FTE_Research_Pct
                c_fr = ws.cell(row=row_idx, column=12, value=r['FTE_Research_Pct'])
                c_fr.number_format = '0%'
                c_fr.alignment = Alignment(horizontal="right")

                # Col M: FTE_LabOps_Pct
                c_fl = ws.cell(row=row_idx, column=13, value=r['FTE_LabOps_Pct'])
                c_fl.number_format = '0%'
                c_fl.alignment = Alignment(horizontal="right")

                # Col N: Total_FTE_Pct (Excel Formula: =SUM(J{row}:M{row}))
                c_ftot = ws.cell(row=row_idx, column=14, value=f"=SUM(J{row_idx}:M{row_idx})")
                c_ftot.number_format = '0%'
                c_ftot.font = bold_font
                c_ftot.alignment = Alignment(horizontal="right")

                # Col O: Notes
                ws.cell(row=row_idx, column=15, value=r['Department_Head_Notes'])

                # Apply font, border, background
                for c_idx in range(1, len(fieldnames) + 1):
                    cell = ws.cell(row=row_idx, column=c_idx)
                    cell.border = gray_border
                    if not cell.font or cell.font.name != "Segoe UI":
                        cell.font = data_font
                    if is_even:
                        cell.fill = zebra_fill

            # -------------------------------------------------------------
            # 4. Attach In-Cell Data Validation Dropdowns
            # -------------------------------------------------------------
            dv_occupancy = DataValidation(
                type="list",
                formula1='"Filled,Vacant"',
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Invalid Selection",
                error="Please select either 'Filled' or 'Vacant'."
            )
            ws.add_data_validation(dv_occupancy)
            dv_occupancy.add(f"C2:C{max_row + 20}")

            dv_type = DataValidation(
                type="list",
                formula1='"Military,Civilian,MOA_Courtesy"',
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Invalid Selection",
                error="Please select 'Military', 'Civilian', or 'MOA_Courtesy'."
            )
            ws.add_data_validation(dv_type)
            dv_type.add(f"D2:D{max_row + 20}")

            dv_tier = DataValidation(
                type="list",
                formula1='"Line_Faculty,Course_Director,Department_Head,Division_Chief,Lab_Director,Research_Exempt"',
                allow_blank=False,
                showErrorMessage=True,
                errorTitle="Invalid Selection",
                error="Please select a valid tier from the dropdown list."
            )
            ws.add_data_validation(dv_tier)
            dv_tier.add(f"E2:E{max_row + 20}")

            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if val_str.startswith('='):
                        val_str = '100%'
                    max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            wb.save(xlsx_path)
            generated_files[dept_code] = xlsx_path

        return generated_files

