"""
Unit Tests for XLSX Roster Generation, Dropdowns, and Dual-Rule FTE Section Equivalents
--------------------------------------------------------------------------------------
Verifies:
1. Dual-rule mathematical section equivalents computation:
   - Rule 1: Implied Ratio Model (actual_sections > 0 and FTE_Academics_Pct > 0)
   - Rule 2: Baseline Anchor Model (actual_sections == 0 or FTE_Academics_Pct == 0, 25% FTE = 1 Sec)
2. RosterGenerator producing valid .xlsx files with:
   - In-cell DataValidation dropdowns on Billet_Occupancy, Billet_Type, Expected_Tier
   - 4-bucket FTE distribution percentages
   - =SUM(J{row}:M{row}) automated check formula
3. RosterManager parsing both .xlsx and .csv files.
"""

import os
import sys
import shutil
import tempfile
import unittest
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from analyzer.roster_manager import (
    compute_section_equivalents,
    RosterManager,
    RosterEntry,
    parse_tier_to_sections,
)
from analyzer.roster_generator import RosterGenerator


class TestSectionEquivalents(unittest.TestCase):

    def test_rule1_implied_ratio_exact(self):
        """Rule 1: If actual_sections > 0 and FTE_Acad > 0, compute implied capacity."""
        # 2 sections taught with 50% Academics, 25% Admin, 25% Research
        # Implied Capacity = 2.0 / 0.50 = 4.0 sections
        # Admin = 4.0 * 0.25 = 1.0 section
        # Research = 4.0 * 0.25 = 1.0 section
        # Gross = 2.0 + 1.0 + 1.0 = 4.0 sections
        res = compute_section_equivalents(
            actual_sections=2.0,
            fte_acad=0.50,
            fte_admin=0.25,
            fte_res=0.25,
            fte_lab=0.0
        )
        self.assertEqual(res['admin'], 1.0)
        self.assertEqual(res['research'], 1.0)
        self.assertEqual(res['labops'], 0.0)
        self.assertEqual(res['gross_burden'], 4.0)

    def test_rule1_implied_ratio_with_string_percentages(self):
        """Rule 1 handles string inputs like '75%', '10%', etc."""
        # 3.0 sections taught, 75% Acad, 10% Admin, 15% Research
        # Implied capacity = 3.0 / 0.75 = 4.0
        # Admin = 4.0 * 0.10 = 0.40
        # Research = 4.0 * 0.15 = 0.60
        # Gross = 3.0 + 0.40 + 0.60 = 4.00
        res = compute_section_equivalents(
            actual_sections="3.0",
            fte_acad="75%",
            fte_admin="10%",
            fte_res="15%",
            fte_lab="0%"
        )
        self.assertEqual(res['admin'], 0.40)
        self.assertEqual(res['research'], 0.60)
        self.assertEqual(res['labops'], 0.0)
        self.assertEqual(res['gross_burden'], 4.0)

    def test_rule2_reversion_zero_sections(self):
        """Rule 2: If actual_sections == 0, revert to 25% FTE = 1 Section."""
        # A department head or full-time admin teaching 0 sections
        # 0% Academics, 75% Admin, 25% Research
        # Admin = 0.75 / 0.25 = 3.0 sections
        # Research = 0.25 / 0.25 = 1.0 section
        # Gross = 0.0 + 3.0 + 1.0 = 4.0 sections
        res = compute_section_equivalents(
            actual_sections=0.0,
            fte_acad=0.0,
            fte_admin=0.75,
            fte_res=0.25,
            fte_lab=0.0
        )
        self.assertEqual(res['admin'], 3.0)
        self.assertEqual(res['research'], 1.0)
        self.assertEqual(res['labops'], 0.0)
        self.assertEqual(res['gross_burden'], 4.0)

    def test_rule2_reversion_zero_academic_fte(self):
        """Rule 2: Even if actual_sections > 0 (e.g. voluntary overload while 100% Admin), if FTE_Acad == 0, revert to 25% FTE = 1 Sec."""
        res = compute_section_equivalents(
            actual_sections=1.0,
            fte_acad=0.0,
            fte_admin=0.80,
            fte_res=0.20,
            fte_lab=0.0
        )
        # Admin = 0.80 / 0.25 = 3.2 sections
        # Research = 0.20 / 0.25 = 0.8 sections
        # Gross = 1.0 + 3.2 + 0.8 = 5.0 sections
        self.assertEqual(res['admin'], 3.2)
        self.assertEqual(res['research'], 0.8)
        self.assertEqual(res['gross_burden'], 5.0)

    def test_lab_operations_relief(self):
        """Test lab operations percentage relief calculation."""
        # 2 sections taught, 50% Acad, 20% LabOps, 30% Admin
        # Implied capacity = 2.0 / 0.50 = 4.0
        # LabOps = 4.0 * 0.20 = 0.8
        # Admin = 4.0 * 0.30 = 1.2
        # Gross = 2.0 + 0.8 + 1.2 = 4.0
        res = compute_section_equivalents(
            actual_sections=2.0,
            fte_acad=0.50,
            fte_admin=0.30,
            fte_res=0.0,
            fte_lab=0.20
        )
        self.assertEqual(res['labops'], 0.8)
        self.assertEqual(res['admin'], 1.2)
        self.assertEqual(res['gross_burden'], 4.0)


class TestRosterGeneratorAndManager(unittest.TestCase):

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_roster_generator_outputs_valid_xlsx(self):
        """Verify RosterGenerator produces valid .xlsx files with dropdown validations and formulas."""
        mock_metrics = {
            'departments': [
                {'dept_code': 'ESME', 'dept_name': 'Mechanical Engineering'}
            ],
            'faculty_directory': [
                {
                    'instructor': 'TestProf, Alpha',
                    'primary_dept': 'ESME',
                    'weighted_sections': 3.0,
                    'cadet_load_allocated': 60.0,
                    'advisees_count': 15,
                    'courses_taught': ['MECHENG 312']
                },
                {
                    'instructor': 'DeptHead, Beta',
                    'primary_dept': 'ESME',
                    'weighted_sections': 1.0,
                    'cadet_load_allocated': 18.0,
                    'advisees_count': 5,
                    'courses_taught': ['MECHENG 499']
                }
            ]
        }

        gen = RosterGenerator(mock_metrics, output_dir=self.temp_dir)
        files = gen.generate_all_rosters()

        self.assertIn('ESME', files)
        xlsx_path = files['ESME']
        self.assertTrue(os.path.exists(xlsx_path))
        self.assertTrue(xlsx_path.endswith('.xlsx'))

        # Also verify backwards-compatible CSV exists
        csv_path = os.path.join(self.temp_dir, "starter_roster_ESME.csv")
        self.assertTrue(os.path.exists(csv_path))

        # Inspect XLSX Workbook
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        self.assertEqual(ws.title, "Roster_ESME")

        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Department_Code", headers)
        self.assertIn("Billet_Occupancy", headers)
        self.assertIn("Billet_Type", headers)
        self.assertIn("Expected_Tier", headers)
        self.assertIn("FTE_Academics_Pct", headers)
        self.assertIn("FTE_Admin_Pct", headers)
        self.assertIn("FTE_Research_Pct", headers)
        self.assertIn("FTE_LabOps_Pct", headers)
        self.assertIn("Total_FTE_Pct", headers)

        # Check formula in Total_FTE_Pct (Col 14 / N)
        # Row 2 (Alpha) should have formula =SUM(J2:M2)
        total_cell = ws.cell(row=2, column=14)
        self.assertEqual(total_cell.value, "=SUM(J2:M2)")

        # Verify DataValidations exist on worksheet
        val_list = ws.data_validations.dataValidation
        self.assertGreaterEqual(len(val_list), 3)

        formula_strings = [v.formula1 for v in val_list]
        self.assertTrue(any('"Filled,Vacant"' in f for f in formula_strings))
        self.assertTrue(any('"Military,Civilian,MOA_Courtesy"' in f for f in formula_strings))
        self.assertTrue(any('Line_Faculty' in f for f in formula_strings))

    def test_roster_manager_ingests_xlsx(self):
        """Verify RosterManager loads and parses generated XLSX files with section equivalents."""
        mock_metrics = {
            'departments': [{'dept_code': 'ESCS', 'dept_name': 'Computer Science'}],
            'faculty_directory': [
                {
                    'instructor': 'Turing, Alan',
                    'primary_dept': 'ESCS',
                    'weighted_sections': 2.0,
                    'cadet_load_allocated': 40.0,
                    'advisees_count': 10,
                    'courses_taught': ['COMPENG 380']
                }
            ]
        }

        gen = RosterGenerator(mock_metrics, output_dir=self.temp_dir)
        files = gen.generate_all_rosters()
        xlsx_path = files['ESCS']

        # Now load with RosterManager
        rm = RosterManager()
        loaded = rm.load_roster_files([xlsx_path])
        self.assertGreaterEqual(loaded, 1)

        entry = rm.get_entry('Turing, Alan')
        self.assertIsNotNone(entry)
        self.assertEqual(entry.department_code, 'ESCS')
        self.assertEqual(entry.billet_occupancy, 'Filled')
        self.assertEqual(entry.billet_type, 'Military')
        self.assertTrue(rm.has_fte_data)


if __name__ == '__main__':
    unittest.main()
